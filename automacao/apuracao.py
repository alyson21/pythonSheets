"""Automação Apuração: insere notas extraídas (SIEG) nas seções de cada cliente.

Fluxo:
  1. Anexa a planilha modelo (APURAÇÃO), que tem uma aba por cliente.
  2. Cada aba tem 3 seções (categorias), cujos títulos variam — são detectadas
     pelo cabeçalho "Download ... Valor" e pela linha TOTAL com =SUM().
  3. Para cada cliente o usuário anexa um arquivo SIEG a cada seção desejada.
  4. As notas são inseridas acima da linha TOTAL, que tem o =SUM() estendido.

O "lado" da seção decide qual parte da nota vai na coluna de contraparte:
  - Emitidas / Vendas / Prestados  → cabeçalho "Tomador"/"Cliente"  → Razão Soc. Dest
  - Serviços Tomados / Entradas    → cabeçalho "Prestador"/"Fornecedor" → Razão Soc. Emit
"""

from __future__ import annotations

import json
import re
import shutil
import threading
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import (
    BOTH,
    DISABLED,
    END,
    GROOVE,
    LEFT,
    NORMAL,
    RIGHT,
    W,
    WORD,
    X,
    Y,
    Button,
    Frame,
    Label,
    Scrollbar,
    StringVar,
    Text,
    filedialog,
    messagebox,
    ttk,
)

import openpyxl

from automacao.caminhos import DADOS
from automacao.excel import copiar_estilo, parse_data, parse_numero, tem_lock

PLANILHA_DIR = DADOS / "planilhaFinal"
PROCESSADOS = DADOS / "processados"
JOURNAL = PROCESSADOS / "apuracao.journal.json"

# Rótulos da coluna de contraparte (col F) que indicam que o cliente é o EMITENTE
# da nota (venda/prestação). Nesses casos a contraparte é o destinatário.
LADO_SAIDA_LABELS = {"tomador", "cliente"}


def planilha_padrao() -> Path:
    """Escolhe automaticamente a planilha de apuração mais recente."""
    candidatos = sorted(PLANILHA_DIR.glob("*.xlsx"))
    apuracao = [p for p in candidatos if "apura" in p.name.casefold()]
    escolha = apuracao or candidatos
    if escolha:
        return max(escolha, key=lambda p: p.stat().st_mtime)
    return PLANILHA_DIR / "APURAÇÃO.xlsx"


# ── detecção das seções no modelo ────────────────────────────────────────────

@dataclass
class Secao:
    titulo: str
    header_row: int
    total_row: int
    label_contraparte: str
    lado: str  # "saida" (contraparte=destinatário) | "entrada" (contraparte=emitente)


def _texto(cell) -> str:
    v = cell.value
    return v.strip() if isinstance(v, str) else ""


def detectar_secoes(ws) -> list[Secao]:
    """Localiza as seções da aba pelo cabeçalho 'Download' e pela linha =SUM()."""
    headers = [r for r in range(1, ws.max_row + 1)
               if _texto(ws.cell(row=r, column=1)) == "Download"]
    secoes: list[Secao] = []

    for idx, header_row in enumerate(headers):
        limite = headers[idx + 1] if idx + 1 < len(headers) else ws.max_row + 1

        # título = primeira célula não vazia da coluna A acima do cabeçalho
        titulo = ""
        for r in range(header_row - 1, 0, -1):
            t = _texto(ws.cell(row=r, column=1))
            if t and t != "TOTAL":
                titulo = t
                break

        # linha TOTAL = primeira linha da seção com =SUM() na coluna G (7)
        total_row = None
        for r in range(header_row + 1, limite):
            g = ws.cell(row=r, column=7).value
            if isinstance(g, str) and g.upper().startswith("=SUM("):
                total_row = r
                break
        if total_row is None:
            continue

        label = _texto(ws.cell(row=header_row, column=6))
        lado = "saida" if label.casefold() in LADO_SAIDA_LABELS else "entrada"
        secoes.append(Secao(titulo or f"Seção {header_row}", header_row, total_row, label, lado))

    return secoes


def titulos_por_cliente(planilha: Path) -> dict[str, list[str]]:
    """Mapeia cada aba (cliente) para a lista de títulos de seção detectados."""
    wb = openpyxl.load_workbook(planilha, read_only=False)
    try:
        return {nome: [s.titulo for s in detectar_secoes(wb[nome])] for nome in wb.sheetnames}
    finally:
        wb.close()


# ── leitura do arquivo SIEG ──────────────────────────────────────────────────

@dataclass
class NotaSieg:
    numero: object
    valor: object
    data: object
    razao_emit: str
    razao_dest: str
    cancelada: bool


def ler_sieg(arquivo: Path) -> list[NotaSieg]:
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    try:
        ws = wb.active
        header_row, cols = None, {}
        for r in range(1, min(ws.max_row, 15) + 1):
            valores = {_texto(ws.cell(row=r, column=c)): c
                       for c in range(1, ws.max_column + 1)
                       if _texto(ws.cell(row=r, column=c))}
            if "Valor" in valores and "Data Emissão" in valores and \
                    any(k.startswith("Num") for k in valores):
                header_row, cols = r, valores
                break
        if header_row is None:
            raise ValueError("cabeçalho SIEG não encontrado (esperado 'Num..., Valor, Data Emissão').")

        col_num = next(c for k, c in cols.items() if k.startswith("Num"))
        col_val = cols["Valor"]
        col_data = cols["Data Emissão"]
        col_emit = cols.get("Razão Soc. Emit")
        col_dest = cols.get("Razão Soc. Dest")
        col_status = cols.get("Status")
        col_tipo = cols.get("Tipo do Evento")

        def get(row, col):
            return ws.cell(row=row, column=col).value if col else None

        notas: list[NotaSieg] = []
        for r in range(header_row + 1, ws.max_row + 1):
            numero = get(r, col_num)
            if numero in (None, ""):
                continue
            status = f"{get(r, col_status) or ''} {get(r, col_tipo) or ''}".casefold()
            notas.append(NotaSieg(
                numero=numero,
                valor=parse_numero(get(r, col_val)),
                data=parse_data(get(r, col_data)),
                razao_emit=str(get(r, col_emit) or "").strip(),
                razao_dest=str(get(r, col_dest) or "").strip(),
                cancelada="cancel" in status,
            ))
        return notas
    finally:
        wb.close()


# ── inserção ─────────────────────────────────────────────────────────────────

_REF_RE = re.compile(r"(?<![A-Za-z0-9_!$])(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _shift_formulas(ws, threshold: int, delta: int) -> None:
    """Desloca em `delta` as referências de linha >= threshold em todas as fórmulas."""
    def repl(m):
        col_abs, col, row_abs, row = m.groups()
        n = int(row)
        if n >= threshold:
            n += delta
        return f"{col_abs}{col}{row_abs}{n}"

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                cell.value = _REF_RE.sub(repl, v)


def _shift_merges(ws, threshold: int, delta: int) -> None:
    """Desloca em `delta` os merges cuja primeira linha é >= threshold.

    O openpyxl move as células ao inserir/remover linhas, mas NÃO reposiciona os
    ranges mesclados, desalinhando títulos e a linha TOTAL das seções seguintes.
    Ajustamos os ranges diretamente (sem unmerge/merge, que apaga células).
    """
    for m in ws.merged_cells.ranges:
        if m.min_row >= threshold:
            m.shift(0, delta)


def _linha_modelo(nota: NotaSieg, lado: str) -> list:
    contraparte = nota.razao_dest if lado == "saida" else nota.razao_emit
    if nota.cancelada:
        return ["Cancelada", nota.numero, None, nota.data, None, contraparte, " -", nota.valor]
    return [None, nota.numero, None, nota.data, None, contraparte, nota.valor, None]


def _linha_vazia(ws, r: int) -> bool:
    return all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, 9))


def _inserir_secao(ws, secao: Secao, notas: list[NotaSieg]) -> tuple[int, int]:
    """Insere as notas com os dados contíguos, logo acima da linha TOTAL.

    Remove as linhas em branco reservadas dentro da seção para não deixar buracos.
    Retorna (primeira_linha_inserida, linhas_em_branco_removidas).
    """
    n = len(notas)
    total_row = secao.total_row

    # estilo de referência = linha logo acima do TOTAL (espaçador já formatado como dado)
    estilos = [ws.cell(row=total_row - 1, column=c) for c in range(1, 9)]
    estilos = [(c.font, c.fill, c.border, c.alignment, c.number_format, c.protection, c.has_style)
               for c in estilos]

    # remove linhas em branco entre o último dado real e o TOTAL
    last_dado = secao.header_row
    for r in range(secao.header_row + 1, total_row):
        if not _linha_vazia(ws, r):
            last_dado = r
    n_blank = total_row - 1 - last_dado
    if n_blank > 0:
        ws.delete_rows(last_dado + 1, n_blank)
        _shift_formulas(ws, total_row, -n_blank)
        _shift_merges(ws, total_row, -n_blank)
        total_row -= n_blank

    insert_at = total_row  # novas linhas entram imediatamente antes do TOTAL
    ws.insert_rows(insert_at, n)
    _shift_formulas(ws, insert_at, n)  # empurra fórmulas abaixo do ponto de inserção
    _shift_merges(ws, insert_at, n)    # e os merges (títulos) das seções seguintes

    for i, nota in enumerate(notas):
        linha = _linha_modelo(nota, secao.lado)
        for col, valor in enumerate(linha, start=1):
            cell = ws.cell(row=insert_at + i, column=col, value=valor)
            f, fill, b, al, nf, prot, has = estilos[col - 1]
            if has:
                cell.font = copy(f)
                cell.fill = copy(fill)
                cell.border = copy(b)
                cell.alignment = copy(al)
                cell.number_format = nf
                cell.protection = copy(prot)

    # estende o =SUM() do TOTAL para cobrir as novas linhas
    novo_total = insert_at + n
    primeiro_dado = secao.header_row + 1
    ws.cell(row=novo_total, column=7).value = f"=SUM(G{primeiro_dado}:G{novo_total - 1})"
    return insert_at, n_blank


def _slug(texto: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", texto).strip("_")[:60] or "x"


def inserir(planilha: Path, fila: list, log) -> None:
    """fila = lista de (cliente: str, {titulo_secao: Path})."""
    PROCESSADOS.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(planilha)
    insercoes = []

    for cliente, mapa in fila:
        if cliente not in wb.sheetnames:
            log(f"  Aba '{cliente}' não existe. Pulado.\n")
            continue

        ws = wb[cliente]
        secoes = {s.titulo: s for s in detectar_secoes(ws)}
        itens = [(titulo, Path(arq)) for titulo, arq in mapa.items() if arq]
        # de baixo para cima: inserir nas seções inferiores não desloca as superiores
        itens.sort(key=lambda t: secoes[t[0]].total_row if t[0] in secoes else 0, reverse=True)

        log(f"▶ Cliente '{cliente}'")
        for titulo, arquivo in itens:
            secao = secoes.get(titulo)
            if secao is None:
                log(f"  Seção '{titulo}' não encontrada. Pulado.")
                continue
            try:
                notas = ler_sieg(arquivo)
            except Exception as e:
                log(f"  ERRO ao ler {arquivo.name}: {e}")
                continue
            if not notas:
                log(f"  '{titulo}': arquivo sem notas. Pulado.")
                continue

            primeira, n_blank = _inserir_secao(ws, secao, notas)
            # de baixo p/ cima: editar esta seção (mais acima) desloca os blocos já
            # registrados desta aba pelo saldo líquido de linhas (inseridas - brancas).
            net_delta = len(notas) - n_blank
            if net_delta:
                for ins in insercoes:
                    if ins["aba"] == cliente:
                        ins["primeira_linha"] += net_delta

            copia = PROCESSADOS / f"{_slug(cliente)}__{_slug(titulo)}.xlsx"
            try:
                shutil.copy2(str(arquivo), str(copia))
            except OSError:
                copia = None
            insercoes.append({
                "aba": cliente,
                "titulo": titulo,
                "primeira_linha": primeira,
                "n": len(notas),
                "bruto": str(copia) if copia else None,
            })
            canceladas = sum(1 for x in notas if x.cancelada)
            extra = f" ({canceladas} cancelada(s))" if canceladas else ""
            log(f"  '{titulo}': {len(notas)} nota(s) inseridas na linha {primeira}{extra}.")
        log("")

    if not insercoes:
        log("Nada foi inserido.\n")
        return

    try:
        wb.save(planilha)
    except PermissionError:
        log("  ERRO: planilha está aberta. Feche e tente novamente.\n")
        return

    _gravar_journal({
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "planilha": str(planilha),
        "insercoes": insercoes,
    })


# ── desfazer ─────────────────────────────────────────────────────────────────

def _ler_journal() -> list:
    if not JOURNAL.exists():
        return []
    try:
        return json.loads(JOURNAL.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []


def _gravar_journal(batch: dict) -> None:
    dados = _ler_journal()
    dados.append(batch)
    JOURNAL.write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")


def desfazer(planilha: Path, log) -> None:
    dados = _ler_journal()
    if not dados:
        log("Nada para desfazer.\n")
        return

    batch = dados[-1]
    wb = openpyxl.load_workbook(planilha)

    # remover de baixo para cima para não invalidar as linhas ainda não removidas
    insercoes = sorted(batch["insercoes"], key=lambda x: x["primeira_linha"], reverse=True)
    for ins in insercoes:
        aba, primeira, n = ins["aba"], ins["primeira_linha"], ins["n"]
        if aba not in wb.sheetnames:
            log(f"  Aba '{aba}' não existe. Pulado.")
            continue
        ws = wb[aba]

        total_row = primeira + n
        formula = ws.cell(row=total_row, column=7).value
        m = re.match(r"=SUM\(G(\d+):G\d+\)", str(formula or ""))
        primeiro_dado = int(m.group(1)) if m else None

        ws.delete_rows(primeira, n)
        _shift_formulas(ws, primeira + n, -n)  # puxa fórmulas de baixo para cima
        _shift_merges(ws, primeira + n, -n)    # e os merges (títulos) das seções seguintes
        if primeiro_dado is not None:
            novo_total = total_row - n
            ws.cell(row=novo_total, column=7).value = f"=SUM(G{primeiro_dado}:G{novo_total - 1})"
        log(f"↩ '{ins['titulo']}' em '{aba}': {n} linha(s) removida(s).")

    try:
        wb.save(planilha)
    except PermissionError:
        log("  ERRO: planilha está aberta. Feche e tente novamente.\n")
        return

    for ins in batch["insercoes"]:
        bruto = ins.get("bruto")
        if bruto and Path(bruto).exists():
            Path(bruto).unlink()

    dados.pop()
    if dados:
        JOURNAL.write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")
    else:
        JOURNAL.unlink(missing_ok=True)
    log("── Desfazer concluído ──\n")


# ── interface gráfica ────────────────────────────────────────────────────────

class ApuracaoPanel(Frame):
    """Insere notas SIEG nas seções (categorias) de cada cliente da apuração."""

    NOME = "Apuração"

    def __init__(self, master):
        super().__init__(master, padx=12, pady=10)
        self.clientes: dict[str, list[str]] = {}
        self.slot_vars: list[tuple[str, StringVar]] = []   # [(titulo, StringVar(path))]
        self.fila: dict[str, dict[str, str]] = {}          # cliente -> {titulo: path}
        self._build()

    # ── construção ──
    def _build(self):
        P = dict(padx=6, pady=4)

        fr_pf = Frame(self, relief=GROOVE, bd=1)
        fr_pf.pack(fill=X, **P)
        Label(fr_pf, text="Planilha de Apuração (modelo)", font=("", 9, "bold")).pack(
            anchor=W, padx=6, pady=(4, 0))
        row_pf = Frame(fr_pf)
        row_pf.pack(fill=X, padx=6, pady=4)
        self.var_pf = StringVar(value=str(planilha_padrao()))
        ttk.Entry(row_pf, textvariable=self.var_pf, width=54).pack(side=LEFT)
        Button(row_pf, text="Selecionar", command=self._sel_pf).pack(side=LEFT, padx=(6, 0))
        Button(row_pf, text="Carregar", command=self._carregar).pack(side=LEFT, padx=(6, 0))

        fr_cli = Frame(self, relief=GROOVE, bd=1)
        fr_cli.pack(fill=X, **P)
        Label(fr_cli, text="Cliente (aba)", font=("", 9, "bold")).pack(
            anchor=W, padx=6, pady=(4, 0))
        row_cli = Frame(fr_cli)
        row_cli.pack(fill=X, padx=6, pady=4)
        self.var_cli = StringVar()
        self.cb_cli = ttk.Combobox(row_cli, textvariable=self.var_cli, width=50, state="disabled")
        self.cb_cli.pack(side=LEFT)
        self.cb_cli.bind("<<ComboboxSelected>>", lambda e: self._montar_slots())

        self.fr_slots = Frame(fr_cli)
        self.fr_slots.pack(fill=X, padx=6, pady=(0, 4))

        self.btn_add = Button(fr_cli, text="＋ Adicionar cliente à fila",
                              command=self._add_fila, state=DISABLED)
        self.btn_add.pack(anchor=W, padx=6, pady=(0, 6))

        fr_fila = Frame(self, relief=GROOVE, bd=1)
        fr_fila.pack(fill=X, **P)
        Label(fr_fila, text="Fila de processamento", font=("", 9, "bold")).pack(
            anchor=W, padx=6, pady=(4, 0))
        self.fr_lista = Frame(fr_fila)
        self.fr_lista.pack(fill=X, padx=6, pady=(0, 4))

        fr_btn = Frame(self)
        fr_btn.pack(**P)
        self.btn_exec = Button(fr_btn, text="▶  Executar", width=18, font=("", 10, "bold"),
                               bg="#4CAF50", fg="white", activebackground="#45a049",
                               command=self._executar)
        self.btn_exec.pack(side=LEFT, padx=8)
        self.btn_desfaz = Button(fr_btn, text="↩  Desfazer", width=18, font=("", 10, "bold"),
                                 bg="#e53935", fg="white", activebackground="#c62828",
                                 command=self._desfazer)
        self.btn_desfaz.pack(side=LEFT, padx=8)

        fr_log = Frame(self, relief=GROOVE, bd=1)
        fr_log.pack(fill=BOTH, expand=True, **P)
        Label(fr_log, text="Log", font=("", 9, "bold")).pack(anchor=W, padx=6, pady=(4, 0))
        fr_txt = Frame(fr_log)
        fr_txt.pack(fill=BOTH, expand=True, padx=6, pady=4)
        sb = Scrollbar(fr_txt)
        sb.pack(side=RIGHT, fill=Y)
        self.log = Text(fr_txt, height=9, font=("Courier", 9),
                        yscrollcommand=sb.set, wrap=WORD, state=DISABLED)
        self.log.pack(side=LEFT, fill=BOTH, expand=True)
        sb.config(command=self.log.yview)

    # ── helpers ──
    def _append_log(self, msg):
        self.log.configure(state=NORMAL)
        self.log.insert(END, msg + "\n")
        self.log.see(END)
        self.log.configure(state=DISABLED)
        self.log.update_idletasks()

    def _set_btns(self, enabled: bool):
        state = NORMAL if enabled else DISABLED
        self.btn_exec.configure(state=state)
        self.btn_desfaz.configure(state=state)
        self.btn_add.configure(state=state if self.clientes else DISABLED)

    def _verificar_planilha(self):
        pf = Path(self.var_pf.get().strip())
        if not pf.exists():
            messagebox.showerror("Erro", f"Planilha não encontrada:\n{pf}")
            return None
        if tem_lock(pf):
            messagebox.showerror("Arquivo aberto",
                                 f"Feche '{pf.name}' no LibreOffice/Excel antes de continuar.")
            return None
        return pf

    # ── planilha / clientes ──
    def _sel_pf(self):
        p = filedialog.askopenfilename(initialdir=str(PLANILHA_DIR),
                                       filetypes=[("Excel", "*.xlsx")])
        if p:
            self.var_pf.set(p)

    def _carregar(self):
        pf = self._verificar_planilha()
        if not pf:
            return
        try:
            self.clientes = titulos_por_cliente(pf)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível ler a planilha:\n{e}")
            return
        self.cb_cli.configure(values=list(self.clientes), state="readonly")
        self.btn_add.configure(state=NORMAL)
        self._append_log(f"Planilha carregada: {len(self.clientes)} cliente(s).")

    def _montar_slots(self):
        for w in self.fr_slots.winfo_children():
            w.destroy()
        self.slot_vars = []
        cliente = self.var_cli.get()
        titulos = self.clientes.get(cliente, [])
        salvos = self.fila.get(cliente, {})
        for titulo in titulos:
            row = Frame(self.fr_slots)
            row.pack(fill=X, pady=1)
            Label(row, text=titulo, width=30, anchor=W).pack(side=LEFT)
            var = StringVar(value=salvos.get(titulo, ""))
            Label(row, textvariable=var, width=34, anchor=W, fg="#1565C0").pack(side=LEFT)
            Button(row, text="Anexar",
                   command=lambda v=var: self._anexar(v)).pack(side=LEFT, padx=(4, 0))
            Button(row, text="✕",
                   command=lambda v=var: v.set("")).pack(side=LEFT, padx=(2, 0))
            self.slot_vars.append((titulo, var))

    def _anexar(self, var: StringVar):
        p = filedialog.askopenfilename(initialdir=str(DADOS / "brutos"),
                                       filetypes=[("Excel", "*.xlsx")])
        if p:
            var.set(p)

    def _add_fila(self):
        cliente = self.var_cli.get()
        if not cliente:
            return
        mapa = {titulo: var.get() for titulo, var in self.slot_vars if var.get().strip()}
        if not mapa:
            messagebox.showwarning("Aviso", "Anexe ao menos um arquivo antes de adicionar à fila.")
            return
        self.fila[cliente] = mapa
        self._render_fila()

    def _render_fila(self):
        for w in self.fr_lista.winfo_children():
            w.destroy()
        if not self.fila:
            Label(self.fr_lista, text="(vazia)", fg="gray").pack(anchor=W)
            return
        for cliente, mapa in self.fila.items():
            row = Frame(self.fr_lista)
            row.pack(fill=X, pady=1)
            resumo = ", ".join(f"{t} ({Path(p).name})" for t, p in mapa.items())
            Label(row, text=f"{cliente}: {resumo}", anchor=W, justify=LEFT).pack(
                side=LEFT, fill=X, expand=True)
            Button(row, text="✕",
                   command=lambda c=cliente: self._rm_fila(c)).pack(side=RIGHT)

    def _rm_fila(self, cliente):
        self.fila.pop(cliente, None)
        self._render_fila()

    # ── ações ──
    def _executar(self):
        pf = self._verificar_planilha()
        if not pf:
            return
        if not self.fila:
            messagebox.showwarning("Aviso", "A fila está vazia.")
            return

        fila = [(cliente, dict(mapa)) for cliente, mapa in self.fila.items()]
        self._set_btns(False)
        self._append_log("── Iniciando inserção ──")

        def _work():
            try:
                inserir(pf, fila, self._append_log)
                self._append_log("── Concluído ──\n")
                self.fila.clear()
                self.log.after(0, self._render_fila)
            except Exception as e:
                self._append_log(f"ERRO: {e}\n")
            finally:
                self.log.after(0, lambda: self._set_btns(True))

        threading.Thread(target=_work, daemon=True).start()

    def _desfazer(self):
        pf = self._verificar_planilha()
        if not pf:
            return
        self._set_btns(False)
        self._append_log("── Desfazendo última inserção ──")

        def _work():
            try:
                desfazer(pf, self._append_log)
            except Exception as e:
                self._append_log(f"ERRO: {e}\n")
            finally:
                self.log.after(0, lambda: self._set_btns(True))

        threading.Thread(target=_work, daemon=True).start()

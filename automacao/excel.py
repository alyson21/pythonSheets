"""Utilitários compartilhados para manipulação de planilhas Excel."""

import re
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.utils.cell import get_column_letter, quote_sheetname, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation


def parse_data(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    return datetime.strptime(str(v).strip(), "%d/%m/%Y")


def parse_numero(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    t = str(v).strip().replace(".", "").replace(",", ".")
    return float(t) if t not in ("", "-") else None


def parse_documento(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v).strip())
    except ValueError:
        return v


def copiar_estilo(origem, destino):
    if not origem.has_style:
        return
    destino.font = copy(origem.font)
    destino.fill = copy(origem.fill)
    destino.border = copy(origem.border)
    destino.alignment = copy(origem.alignment)
    destino.number_format = origem.number_format
    destino.protection = copy(origem.protection)


def aplicar_negrito(cell):
    f = copy(cell.font)
    f.bold = True
    cell.font = f


def estender_formatacao_condicional(ws, nova_ultima_linha: int):
    """Estende ranges de formatação condicional que terminam antes de nova_ultima_linha."""
    regras = list(ws.conditional_formatting._cf_rules.items())
    ws.conditional_formatting = ConditionalFormattingList()
    for cf_obj, rules in regras:
        sqref = str(cf_obj.sqref)
        novo_sqref = re.sub(
            r"([A-Z]+)(\d+):([A-Z]+)(\d+)",
            lambda m: (
                f"{m.group(1)}{m.group(2)}:{m.group(3)}{nova_ultima_linha}"
                if int(m.group(4)) < nova_ultima_linha else m.group(0)
            ),
            sqref,
        )
        for rule in rules:
            ws.conditional_formatting.add(novo_sqref, rule)


def _sqref_toca_coluna(sqref, coluna: int) -> bool:
    for ref in str(sqref).split():
        try:
            min_col, _, max_col, _ = range_boundaries(ref)
        except ValueError:
            continue
        if min_col <= coluna <= max_col:
            return True
    return False


def _formula_lista_classificacao(wb, nome_aba_lista: str = "Lista") -> str | None:
    if nome_aba_lista not in wb.sheetnames:
        return None

    ws_lista = wb[nome_aba_lista]
    ultima = 0
    for row in range(1, ws_lista.max_row + 1):
        if ws_lista.cell(row=row, column=1).value not in (None, ""):
            ultima = row

    if ultima == 0:
        return None

    return f"{quote_sheetname(nome_aba_lista)}!$A$1:$A${ultima}"


def garantir_validacao_classificacao(ws, primeira_linha: int, ultima_linha: int) -> bool:
    """Garante dropdown da coluna C usando a lista de classificações da aba Lista."""
    if primeira_linha > ultima_linha:
        return False

    coluna = 3
    col = get_column_letter(coluna)
    alvo = f"{col}{primeira_linha}:{col}{ultima_linha}"

    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and _sqref_toca_coluna(dv.sqref, coluna):
            dv.add(alvo)
            return True

    formula = _formula_lista_classificacao(ws.parent)
    if not formula:
        return False

    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(alvo)
    return True


def tem_lock(arquivo: Path) -> bool:
    """Retorna True se o .xlsx parece estar aberto em outro programa."""
    lock = arquivo.parent / f".~lock.{arquivo.name}#"
    if lock.exists():
        return True
    try:
        arquivo.open("r+b").close()
    except PermissionError:
        return True
    return False

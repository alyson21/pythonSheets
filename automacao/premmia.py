"""Automação Premmia: regras de negócio e tela."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import threading
from tkinter import (
    BOTH,
    DISABLED,
    END,
    GROOVE,
    LEFT,
    NORMAL,
    RIGHT,
    W,
    WORD,
    X,
    Y,
    Button,
    Entry,
    Frame,
    Label,
    Scrollbar,
    StringVar,
    Text,
    filedialog,
    messagebox,
)
from typing import Callable, Iterable, NamedTuple

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from automacao.caminhos import DADOS

BASE_DIR = DADOS / "automacoesToDo"
BRUTOS_PADRAO = BASE_DIR / "brutos"
MODELO_DIR = BASE_DIR / "modeloFinal"
SAIDA_PADRAO = BASE_DIR / "formatados"

ABA_CONSOLIDADO = "Pagamento Consolidado"
ABA_DETALHADO = "Pagamento Detalhado"

HEADERS_CONSOLIDADO = [
    "Data da transação",
    "CNPJ",
    "Condição de recebimento",
    "Data Prevista do Crédito",
    "Número de transações",
    "Valor Bruto",
    "Valor Desconto",
    "Valor Líquido",
]

CONDICAO_D01 = "D01"
CONDICAO_LINHA_FINAL = "Até o 5º dia útil do mês seguinte"
MOEDA_BR = '_-[$R$-416]\\ * #,##0.00_-;\\-[$R$-416]\\ * #,##0.00_-;_-[$R$-416]\\ * "-"??_-;_-@_-'


@dataclass
class WorkbookData:
    sheets: dict[str, list[list[object]]]


class CellFormat(NamedTuple):
    style: object
    number_format: str
    alignment: object


def modelo_padrao() -> Path:
    modelos = sorted(MODELO_DIR.glob("*.xlsx"))
    if modelos:
        return modelos[0]
    return MODELO_DIR / "05-2026 - 2000.xlsx"


def processar(arquivos: Iterable[Path], modelo: Path, pasta_saida: Path, log: Callable[[str], None]) -> None:
    modelo = Path(modelo)
    pasta_saida = Path(pasta_saida)

    if not modelo.exists():
        raise FileNotFoundError(f"Modelo final não encontrado: {modelo}")

    pasta_saida.mkdir(parents=True, exist_ok=True)

    for arquivo in arquivos:
        arquivo = Path(arquivo)
        log(f"▶ {arquivo.name}")
        try:
            dados = ler_workbook(arquivo)
            destino = pasta_saida / f"{arquivo.stem}.xlsx"
            gerar_saida(dados, modelo, destino)
            log(f"  OK: gerado {destino.name}\n")
        except Exception as exc:
            log(f"  ERRO: {exc}\n")


def ler_workbook(arquivo: Path) -> WorkbookData:
    suffix = arquivo.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _ler_xlsx(arquivo)
    if suffix == ".xls":
        return _ler_xls(arquivo)
    raise ValueError(f"formato não suportado: {arquivo.suffix}")


def gerar_saida(dados: WorkbookData, modelo: Path, destino: Path) -> None:
    linhas_consolidado = _sheet_obrigatoria(dados, ABA_CONSOLIDADO)
    linhas_detalhado = _sheet_obrigatoria(dados, ABA_DETALHADO)
    _validar_headers(linhas_consolidado, HEADERS_CONSOLIDADO, ABA_CONSOLIDADO)

    wb = openpyxl.load_workbook(modelo)

    if ABA_CONSOLIDADO not in wb.sheetnames:
        raise ValueError(f"modelo não possui a aba '{ABA_CONSOLIDADO}'")
    if ABA_DETALHADO not in wb.sheetnames:
        raise ValueError(f"modelo não possui a aba '{ABA_DETALHADO}'")

    _preencher_consolidado(wb[ABA_CONSOLIDADO], linhas_consolidado)
    _preencher_detalhado(wb[ABA_DETALHADO], linhas_detalhado)

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)


def _ler_xlsx(arquivo: Path) -> WorkbookData:
    wb = openpyxl.load_workbook(arquivo, data_only=True, read_only=True)
    try:
        sheets = {}
        for ws in wb.worksheets:
            sheets[ws.title] = _normalizar_linhas(ws.iter_rows(values_only=True))
        return WorkbookData(sheets=sheets)
    finally:
        wb.close()


def _ler_xls(arquivo: Path) -> WorkbookData:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("dependência xlrd não instalada; rode GerarExe.bat novamente") from exc

    book = xlrd.open_workbook(str(arquivo))
    sheets = {}
    for sheet in book.sheets():
        linhas = []
        for row_idx in range(sheet.nrows):
            linha = [
                _valor_xlrd(sheet.cell(row_idx, col_idx), book.datemode)
                for col_idx in range(sheet.ncols)
            ]
            linhas.append(linha)
        sheets[sheet.name] = _normalizar_linhas(linhas)
    return WorkbookData(sheets=sheets)


def _valor_xlrd(cell, datemode: int):
    import xlrd

    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        dt = xlrd.xldate.xldate_as_datetime(cell.value, datemode)
        if dt.time().hour == dt.time().minute == dt.time().second == 0:
            return dt.strftime("%d/%m/%Y")
        return dt.strftime("%d/%m/%Y %H:%M:%S")
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        if float(cell.value).is_integer():
            return int(cell.value)
        return cell.value
    return cell.value


def _normalizar_linhas(linhas: Iterable[Iterable[object]]) -> list[list[object]]:
    normalizadas = []
    for linha in linhas:
        valores = [_normalizar_valor(v) for v in linha]
        while valores and valores[-1] in (None, ""):
            valores.pop()
        if any(v not in (None, "") for v in valores):
            normalizadas.append(valores)
    return normalizadas


def _normalizar_valor(valor):
    if isinstance(valor, str):
        valor = valor.strip()
        return valor if valor != "" else None
    if isinstance(valor, datetime):
        if valor.time().hour == valor.time().minute == valor.time().second == 0:
            return valor.strftime("%d/%m/%Y")
        return valor.strftime("%d/%m/%Y %H:%M:%S")
    return valor


def _sheet_obrigatoria(dados: WorkbookData, nome: str) -> list[list[object]]:
    for sheet_name, linhas in dados.sheets.items():
        if sheet_name.strip().casefold() == nome.casefold():
            return linhas
    raise ValueError(f"aba obrigatória não encontrada: {nome}")


def _validar_headers(linhas: list[list[object]], headers: list[str], nome_aba: str) -> None:
    if not linhas:
        raise ValueError(f"aba '{nome_aba}' está vazia")

    encontrados = [str(v).strip() if v is not None else "" for v in (linhas[0] + [None] * len(headers))[:len(headers)]]
    if encontrados != headers:
        raise ValueError(
            f"cabeçalhos inválidos em '{nome_aba}'. Esperado: {', '.join(headers)}"
        )


def _preencher_consolidado(ws: Worksheet, linhas_origem: list[list[object]]) -> None:
    header = linhas_origem[0][:8]
    linhas = [(list(linha) + [None] * 8)[:8] for linha in linhas_origem[1:]]
    linhas = [linha for linha in linhas if any(v not in (None, "") for v in linha)]

    linhas_sem_d01 = [linha for linha in linhas if _condicao(linha).upper() != CONDICAO_D01]
    linhas_finais = [
        linha for linha in linhas_sem_d01
        if _eh_linha_final(linha)
    ]
    linhas_dados = [
        linha for linha in linhas_sem_d01
        if not _eh_linha_final(linha)
    ]
    linhas_dados.sort(key=lambda linha: _condicao(linha), reverse=True)

    total_row = 2 + len(linhas_dados)
    final_row = total_row + 8
    formatos = _capturar_formatos_consolidado(ws)
    _ajustar_quantidade_linhas(ws, final_row)
    _limpar_valores(ws, final_row, 8)

    _escrever_linha(ws, 1, header)
    _aplicar_formato_linha(ws, 1, formatos["header"])
    for idx, linha in enumerate(linhas_dados, start=2):
        _escrever_linha(ws, idx, linha)
        _aplicar_formato_linha(ws, idx, formatos["dados"])

    if linhas_dados:
        ws.cell(total_row, 6, f"=SUM(F2:F{total_row - 1})")
        ws.cell(total_row, 7, f"=SUM(G2:G{total_row - 1})")
        ws.cell(total_row, 8, f"=SUM(H2:H{total_row - 1})")
    _aplicar_formato_linha(ws, total_row, formatos["total"])

    for row in range(total_row + 1, final_row):
        _aplicar_formato_linha(ws, row, formatos["vazio"])

    if linhas_finais:
        _escrever_linha(ws, final_row, linhas_finais[0])
    _aplicar_formato_linha(ws, final_row, formatos["final"])

    _formatar_consolidado(ws, total_row, final_row)


def _preencher_detalhado(ws: Worksheet, linhas_origem: list[list[object]]) -> None:
    max_col = max((len(linha) for linha in linhas_origem), default=1)
    header_fmt = _capturar_formato_linha(ws, 1, max_col)
    data_fmt = _capturar_formato_linha(ws, 2 if ws.max_row >= 2 else 1, max_col)
    _ajustar_quantidade_linhas(ws, len(linhas_origem))
    _limpar_valores(ws, len(linhas_origem), max(ws.max_column, max_col))

    for row_idx, linha in enumerate(linhas_origem, start=1):
        _escrever_linha(ws, row_idx, linha)
        _aplicar_formato_linha(ws, row_idx, header_fmt if row_idx == 1 else data_fmt)

    if linhas_origem:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{len(linhas_origem)}"
        _aplicar_bordas(ws, 1, len(linhas_origem), 1, max_col)


def _condicao(linha: list[object]) -> str:
    valor = linha[2] if len(linha) > 2 else ""
    return str(valor or "").strip()


def _eh_linha_final(linha: list[object]) -> bool:
    condicao = _condicao(linha).casefold().replace("°", "º")
    return condicao == CONDICAO_LINHA_FINAL.casefold()


def _ajustar_quantidade_linhas(ws: Worksheet, total_linhas: int) -> None:
    if ws.max_row > total_linhas:
        ws.delete_rows(total_linhas + 1, ws.max_row - total_linhas)


def _limpar_valores(ws: Worksheet, max_row: int, max_col: int) -> None:
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None


def _escrever_linha(ws: Worksheet, row_idx: int, valores: list[object]) -> None:
    for col_idx, valor in enumerate(valores, start=1):
        ws.cell(row=row_idx, column=col_idx, value=valor)


def _formatar_consolidado(ws: Worksheet, total_row: int, final_row: int) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, 9):
        cell = ws.cell(1, col)
        font = copy(cell.font)
        font.bold = True
        cell.font = font
        cell.border = border
        alignment = copy(cell.alignment)
        alignment.horizontal = "center"
        cell.alignment = alignment

    for col in range(6, 9):
        for row in range(2, final_row + 1):
            ws.cell(row, col).number_format = MOEDA_BR

    if total_row > 1:
        _aplicar_bordas(ws, 1, total_row, 1, 8)
    _aplicar_bordas(ws, final_row, final_row, 1, 8)

    for col in range(6, 9):
        cell = ws.cell(total_row, col)
        font = copy(cell.font)
        font.bold = True
        cell.font = font

    ws.auto_filter.ref = f"A1:H{final_row}"


def _capturar_formatos_consolidado(ws: Worksheet) -> dict[str, list[CellFormat]]:
    total_modelo = _linha_com_formula_total(ws) or min(ws.max_row, 2)
    final_modelo = _linha_final_modelo(ws) or ws.max_row
    vazio_modelo = total_modelo + 1 if total_modelo + 1 < final_modelo else total_modelo
    return {
        "header": _capturar_formato_linha(ws, 1, 8),
        "dados": _capturar_formato_linha(ws, 2 if ws.max_row >= 2 else 1, 8),
        "total": _capturar_formato_linha(ws, total_modelo, 8),
        "vazio": _capturar_formato_linha(ws, vazio_modelo, 8),
        "final": _capturar_formato_linha(ws, final_modelo, 8),
    }


def _linha_com_formula_total(ws: Worksheet) -> int | None:
    for row in range(1, ws.max_row + 1):
        valores = [ws.cell(row, col).value for col in range(6, 9)]
        if any(isinstance(valor, str) and valor.startswith("=") for valor in valores):
            return row
    return None


def _linha_final_modelo(ws: Worksheet) -> int | None:
    for row in range(1, ws.max_row + 1):
        if _eh_linha_final([ws.cell(row, col).value for col in range(1, 9)]):
            return row
    return None


def _capturar_formato_linha(ws: Worksheet, row: int, max_col: int) -> list[CellFormat]:
    return [_capturar_formato(ws.cell(row=row, column=col)) for col in range(1, max_col + 1)]


def _capturar_formato(cell) -> CellFormat:
    return CellFormat(copy(cell._style), cell.number_format, copy(cell.alignment))


def _aplicar_formato_linha(ws: Worksheet, row: int, formatos: list[CellFormat]) -> None:
    for col, formato in enumerate(formatos, start=1):
        cell = ws.cell(row=row, column=col)
        cell._style = copy(formato.style)
        cell.number_format = formato.number_format
        cell.alignment = copy(formato.alignment)


def _aplicar_bordas(ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).border = border


# Interface gráfica ---------------------------------------------------------

class PremmiaPanel(Frame):
    """Gera uma planilha formatada Premmia para cada arquivo bruto."""

    NOME = "Premmia"

    def __init__(self, master):
        super().__init__(master, padx=12, pady=10)
        self._build()

    def _build(self):
        P = dict(padx=6, pady=4)

        fr_saida = Frame(self, relief=GROOVE, bd=1)
        fr_saida.pack(fill=X, **P)
        Label(fr_saida, text="Pasta de Saída", font=("", 9, "bold")).pack(
            anchor=W, padx=6, pady=(4, 0))

        row_saida = Frame(fr_saida)
        row_saida.pack(fill=X, padx=6, pady=4)
        self.var_saida = StringVar(value=str(SAIDA_PADRAO))
        Entry(row_saida, textvariable=self.var_saida, width=62).pack(side=LEFT)
        Button(row_saida, text="Selecionar", command=self._sel_saida).pack(
            side=LEFT, padx=(6, 0))

        fr_brutos = Frame(self, relief=GROOVE, bd=1)
        fr_brutos.pack(fill=X, **P)
        Label(fr_brutos, text="Planilhas Brutas", font=("", 9, "bold")).pack(
            anchor=W, padx=6, pady=(4, 0))
        Button(fr_brutos, text="+ Adicionar arquivo(s)", command=self._add_arqs).pack(
            anchor=W, padx=6, pady=4)

        hdr = Frame(fr_brutos)
        hdr.pack(fill=X, padx=6)
        Label(hdr, text="Arquivo", width=58, anchor=W, fg="gray").grid(row=0, column=0)

        self.fr_lista = Frame(fr_brutos)
        self.fr_lista.pack(fill=X, padx=6, pady=(0, 4))
        self.entradas = []
        self._next_row = 0

        fr_btn = Frame(self)
        fr_btn.pack(**P)
        self.btn_exec = Button(
            fr_btn, text="▶  Executar", width=18, font=("", 10, "bold"),
            bg="#4CAF50", fg="white", activebackground="#45a049",
            command=self._executar)
        self.btn_exec.pack(side=LEFT, padx=8)

        fr_log = Frame(self, relief=GROOVE, bd=1)
        fr_log.pack(fill=BOTH, expand=True, **P)
        Label(fr_log, text="Log", font=("", 9, "bold")).pack(
            anchor=W, padx=6, pady=(4, 0))

        fr_txt = Frame(fr_log)
        fr_txt.pack(fill=BOTH, expand=True, padx=6, pady=4)
        sb = Scrollbar(fr_txt)
        sb.pack(side=RIGHT, fill=Y)
        self.log = Text(fr_txt, height=10, font=("Courier", 9),
                        yscrollcommand=sb.set, wrap=WORD, state=DISABLED)
        self.log.pack(side=LEFT, fill=BOTH, expand=True)
        sb.config(command=self.log.yview)

    def _append_log(self, msg):
        self.log.configure(state=NORMAL)
        self.log.insert(END, msg + "\n")
        self.log.see(END)
        self.log.configure(state=DISABLED)
        self.log.update_idletasks()

    def _set_btns(self, enabled: bool):
        self.btn_exec.configure(state=NORMAL if enabled else DISABLED)

    def _sel_saida(self):
        p = filedialog.askdirectory(initialdir=self.var_saida.get().strip() or str(SAIDA_PADRAO))
        if p:
            self.var_saida.set(p)

    def _add_arqs(self):
        paths = filedialog.askopenfilenames(
            initialdir=str(BRUTOS_PADRAO),
            filetypes=[
                ("Excel", "*.xls *.xlsx *.xlsm *.XLS *.XLSX *.XLSM"),
                ("Todos os arquivos", "*.*"),
            ],
        )
        for p in paths:
            path = Path(p)
            if path not in self.entradas:
                self._add_linha(path)

    def _add_linha(self, path: Path):
        row = Frame(self.fr_lista)
        row.grid(row=self._next_row, column=0, sticky=W, pady=1)
        self._next_row += 1

        Label(row, text=path.name, width=58, anchor=W).grid(row=0, column=0)
        Button(row, text="✕", command=lambda: self._rm_linha(path, row)).grid(
            row=0, column=1, padx=(4, 0))
        self.entradas.append(path)

    def _rm_linha(self, path, row):
        self.entradas = [p for p in self.entradas if p != path]
        row.destroy()

    def _executar(self):
        modelo = modelo_padrao()
        saida = Path(self.var_saida.get().strip())

        if not modelo.exists():
            messagebox.showerror(
                "Erro",
                f"Modelo final não encontrado:\n{modelo}\n\n"
                "Verifique a pasta dados/automacoesToDo/modeloFinal.",
            )
            return
        if not self.entradas:
            messagebox.showwarning("Aviso", "Nenhuma planilha bruta adicionada.")
            return

        self._set_btns(False)
        self._append_log("── Iniciando Premmia ──")

        def _work():
            try:
                processar(self.entradas, modelo, saida, self._append_log)
                self._append_log("── Concluído ──\n")
            except Exception as exc:
                self._append_log(f"ERRO: {exc}\n")
            finally:
                self.log.after(0, lambda: self._set_btns(True))

        threading.Thread(target=_work, daemon=True).start()

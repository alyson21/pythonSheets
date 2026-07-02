"""Automação Santander Postos: regras de negócio e tela."""

import shutil
import threading
from collections import defaultdict
from pathlib import Path
from tkinter import (
    BOTH,
    DISABLED,
    END,
    GROOVE,
    LEFT,
    NORMAL,
    RIGHT,
    W,
    WORD,
    X,
    Y,
    Button,
    Entry,
    Frame,
    Label,
    Scrollbar,
    StringVar,
    Text,
    filedialog,
    messagebox,
)

import openpyxl

from automacao.excel import (
    aplicar_negrito,
    copiar_estilo,
    estender_formatacao_condicional,
    garantir_validacao_classificacao,
    tem_lock,
    parse_data,
    parse_documento,
    parse_numero,
)
from automacao.caminhos import DADOS

PROCESSADOS      = DADOS / "processados"
PLANILHA_PADRAO  = DADOS / "planilhaFinal" / "Santander - Postos.xlsx"


def inserir(planilha: Path, entradas: list, log):
    """entradas = lista de (arquivo: Path, aba: str)"""
    PROCESSADOS.mkdir(parents=True, exist_ok=True)

    for arquivo, aba in entradas:
        log(f"▶ {arquivo.name}  →  aba '{aba}'")
        try:
            wb_b = openpyxl.load_workbook(arquivo, data_only=True)
            linhas = [
                l for l in wb_b.active.iter_rows(min_row=2, values_only=True)
                if any(c is not None and c != "" for c in l)
            ]
            if not linhas:
                log("  Sem dados. Pulado.\n"); continue

            wb = openpyxl.load_workbook(planilha)
            if aba not in wb.sheetnames:
                log(f"  Aba '{aba}' não existe na planilha. Pulado.\n"); continue

            ws         = wb[aba]
            ref_normal = max(ws.max_row - 1, 1)
            ref_bold   = ws.max_row
            ini        = ws.max_row + 1

            # referência específica para col E em linhas bold (preserva font color indexed)
            ref_bold_e = ref_bold
            for r in range(2, min(ws.max_row, 500)):
                if ws.cell(row=r, column=1).font.bold:
                    fc = ws.cell(row=r, column=5).font.color
                    if fc and fc.type == "indexed":
                        ref_bold_e = r
                        break

            # índices das últimas linhas de cada dia
            linhas_por_dia = defaultdict(list)
            for i, linha in enumerate(linhas):
                data_val = parse_data((list(linha) + [None] * 5)[0])
                linhas_por_dia[data_val].append(i)
            ultimas_do_dia = {indices[-1] for indices in linhas_por_dia.values()}

            for i, linha in enumerate(linhas):
                v = (list(linha) + [None] * 5)[:5]
                nova = [
                    parse_data(v[0]), v[1], None,
                    parse_documento(v[2]), parse_numero(v[3]), parse_numero(v[4]),
                ]
                ref = ref_bold if i in ultimas_do_dia else ref_normal
                for col, val in enumerate(nova, 1):
                    cell = ws.cell(row=ini + i, column=col, value=val)
                    ref_src = ref_bold_e if (col == 5 and i in ultimas_do_dia) else ref
                    copiar_estilo(ws.cell(row=ref_src, column=col), cell)

            for idx in ultimas_do_dia:
                for col in range(1, 7):
                    aplicar_negrito(ws.cell(row=ini + idx, column=col))

            ultima_inserida = ini + len(linhas) - 1
            estender_formatacao_condicional(ws, ultima_inserida)
            garantir_validacao_classificacao(ws, 2, ultima_inserida)

            try:
                wb.save(planilha)
            except PermissionError:
                log("  ERRO: planilha está aberta. Feche e tente novamente.\n")
                continue

            destino = PROCESSADOS / f"{aba}.xlsx"
            if arquivo.resolve() != destino.resolve():
                shutil.copy2(str(arquivo), str(destino))
            log(f"  {len(linhas)} linhas inseridas (linhas {ini}–{ultima_inserida}), "
                f"{len(ultimas_do_dia)} dia(s) destacado(s).\n")

        except Exception as e:
            log(f"  ERRO: {e}\n")


def desfazer(planilha: Path, log):
    if not PROCESSADOS.exists():
        log("Nada para desfazer.\n"); return

    arqs = sorted(p for p in PROCESSADOS.glob("*.xlsx") if not p.name.startswith("~"))
    if not arqs:
        log("Nada em dados/processados para desfazer.\n"); return

    for arq in arqs:
        aba = arq.stem
        log(f"↩ Desfazendo aba '{aba}'...")
        try:
            wb_b = openpyxl.load_workbook(arq, data_only=True, read_only=True)
            n = sum(
                1 for l in wb_b.active.iter_rows(min_row=2, values_only=True)
                if any(c is not None and c != "" for c in l)
            )
            wb_b.close()
            if n == 0:
                log("  Sem linhas. Pulado.\n"); continue

            wb = openpyxl.load_workbook(planilha)
            if aba not in wb.sheetnames:
                log(f"  Aba '{aba}' não existe. Pulado.\n"); continue

            ws       = wb[aba]
            primeira = ws.max_row - n + 1
            if primeira < 2:
                log("  Erro: removeria cabeçalho. Pulado.\n"); continue

            ws.delete_rows(primeira, n)
            try:
                wb.save(planilha)
            except PermissionError:
                log("  ERRO: planilha está aberta. Feche e tente novamente.\n")
                continue
            arq.unlink()
            log(f"  {n} linhas removidas.\n")

        except Exception as e:
            log(f"  ERRO: {e}\n")


# Interface gráfica ---------------------------------------------------------

class SantanderPostosPanel(Frame):
    """Insere extratos do Santander na planilha 'Santander - Postos.xlsx'."""

    NOME = "Santander Postos"

    def __init__(self, master):
        super().__init__(master, padx=12, pady=10)
        self._build()

    # ── construção ──────────────────────────────────────────────────────────

    def _build(self):
        P = dict(padx=6, pady=4)

        fr_pf = Frame(self, relief=GROOVE, bd=1)
        fr_pf.pack(fill=X, **P)
        Label(fr_pf, text="Planilha Final", font=("", 9, "bold")).pack(
            anchor=W, padx=6, pady=(4, 0))

        row_pf = Frame(fr_pf)
        row_pf.pack(fill=X, padx=6, pady=4)
        self.var_pf = StringVar(value=str(PLANILHA_PADRAO))
        Entry(row_pf, textvariable=self.var_pf, width=62).pack(side=LEFT)
        Button(row_pf, text="Selecionar", command=self._sel_pf).pack(side=LEFT, padx=(6, 0))

        fr_ex = Frame(self, relief=GROOVE, bd=1)
        fr_ex.pack(fill=X, **P)
        Label(fr_ex, text="Extratos Brutos", font=("", 9, "bold")).pack(
            anchor=W, padx=6, pady=(4, 0))
        Button(fr_ex, text="+ Adicionar arquivo(s)", command=self._add_arqs).pack(
            anchor=W, padx=6, pady=4)

        hdr = Frame(fr_ex)
        hdr.pack(fill=X, padx=6)
        Label(hdr, text="Arquivo", width=36, anchor=W, fg="gray").grid(row=0, column=0)
        Label(hdr, text="Aba de destino", width=22, anchor=W, fg="gray").grid(row=0, column=1)

        self.fr_lista = Frame(fr_ex)
        self.fr_lista.pack(fill=X, padx=6, pady=(0, 4))
        self.entradas = []     # [(path, StringVar, Frame)]
        self._next_row = 0

        fr_btn = Frame(self)
        fr_btn.pack(**P)
        self.btn_exec = Button(
            fr_btn, text="▶  Executar", width=18, font=("", 10, "bold"),
            bg="#4CAF50", fg="white", activebackground="#45a049",
            command=self._executar)
        self.btn_exec.pack(side=LEFT, padx=8)
        self.btn_desfaz = Button(
            fr_btn, text="↩  Desfazer", width=18, font=("", 10, "bold"),
            bg="#e53935", fg="white", activebackground="#c62828",
            command=self._desfazer)
        self.btn_desfaz.pack(side=LEFT, padx=8)

        fr_log = Frame(self, relief=GROOVE, bd=1)
        fr_log.pack(fill=BOTH, expand=True, **P)
        Label(fr_log, text="Log", font=("", 9, "bold")).pack(
            anchor=W, padx=6, pady=(4, 0))

        fr_txt = Frame(fr_log)
        fr_txt.pack(fill=BOTH, expand=True, padx=6, pady=4)
        sb = Scrollbar(fr_txt)
        sb.pack(side=RIGHT, fill=Y)
        self.log = Text(fr_txt, height=10, font=("Courier", 9),
                        yscrollcommand=sb.set, wrap=WORD, state=DISABLED)
        self.log.pack(side=LEFT, fill=BOTH, expand=True)
        sb.config(command=self.log.yview)

    # ── helpers ──────────────────────────────────────────────────────────────

    def _append_log(self, msg):
        self.log.configure(state=NORMAL)
        self.log.insert(END, msg + "\n")
        self.log.see(END)
        self.log.configure(state=DISABLED)
        self.log.update_idletasks()

    def _set_btns(self, enabled: bool):
        state = NORMAL if enabled else DISABLED
        self.btn_exec.configure(state=state)
        self.btn_desfaz.configure(state=state)

    def _verificar_planilha(self):
        pf = Path(self.var_pf.get().strip())
        if not pf.exists():
            messagebox.showerror("Erro", f"Planilha não encontrada:\n{pf}")
            return None
        if tem_lock(pf):
            messagebox.showerror(
                "Arquivo aberto",
                f"Feche '{pf.name}' no LibreOffice/Excel antes de continuar.")
            return None
        return pf

    # ── seleção / lista ──────────────────────────────────────────────────────

    def _sel_pf(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.var_pf.set(p)

    def _add_arqs(self):
        paths = filedialog.askopenfilenames(filetypes=[("Excel", "*.xlsx")])
        for p in paths:
            self._add_linha(Path(p))

    def _add_linha(self, path: Path):
        row = Frame(self.fr_lista)
        row.grid(row=self._next_row, column=0, sticky=W, pady=1)
        self._next_row += 1

        Label(row, text=path.name, width=36, anchor=W).grid(row=0, column=0)
        var = StringVar(value=path.stem)
        Entry(row, textvariable=var, width=22).grid(row=0, column=1)
        Button(row, text="✕",
               command=lambda: self._rm_linha(path, row)).grid(row=0, column=2, padx=(4, 0))

        self.entradas.append((path, var, row))

    def _rm_linha(self, path, row):
        self.entradas = [(p, v, r) for p, v, r in self.entradas if p != path]
        row.destroy()

    # ── ações ────────────────────────────────────────────────────────────────

    def _executar(self):
        pf = self._verificar_planilha()
        if not pf:
            return
        if not self.entradas:
            messagebox.showwarning("Aviso", "Nenhum extrato adicionado.")
            return

        entradas = [(p, v.get().strip()) for p, v, _ in self.entradas if v.get().strip()]
        self._set_btns(False)
        self._append_log("── Iniciando inserção ──")

        def _work():
            inserir(pf, entradas, self._append_log)
            self._append_log("── Concluído ──\n")
            self.log.after(0, lambda: self._set_btns(True))

        threading.Thread(target=_work, daemon=True).start()

    def _desfazer(self):
        pf = self._verificar_planilha()
        if not pf:
            return

        self._set_btns(False)
        self._append_log("── Desfazendo inserções ──")

        def _work():
            desfazer(pf, self._append_log)
            self._append_log("── Concluído ──\n")
            self.log.after(0, lambda: self._set_btns(True))

        threading.Thread(target=_work, daemon=True).start()
